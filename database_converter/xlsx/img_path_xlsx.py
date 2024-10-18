from openpyxl import load_workbook
import csv, sys, os.path
import re
import codecs
import time

start_time = time.time()

parent_dir = os.path.dirname(os.path.realpath(__file__ + "/../"))
sys.path.append(parent_dir)
import modules_templates as templates

HULLS_OFFSET = 1    #Index offset
TURRETS_OFFSET = 28
GUNS_OFFSET = 50
SEC_OFFSET = 69

DATABASE_NAME = input()
DATABASE_PATH = os.path.join(sys.path[0], DATABASE_NAME)
JSON_PATH = os.path.join(sys.path[0], "imgPaths.json")
PATH_IMG = "./img/"

def as_string(a): return '"'+a.replace('"', '\\"')+'"' if a else '""'

wb = load_workbook(filename=DATABASE_PATH, read_only=True, data_only=True)
ws = wb['Raw Data']

hulls_string = ""
turrets_string = ""
guns_string = ""

i = 2
while ws.cell(HULLS_OFFSET+0, i).value is not None:
    if i != 2:
        hulls_string += ", \n"
    print(ws.cell(HULLS_OFFSET+0, i).value)
    
    name = ws.cell(HULLS_OFFSET+0, i).value\
                                    .replace("?", "")\
                                    .replace(" (!)", "")
    
    hulls_string += '"'+name.replace("\n", "\\n").replace('"', '\\"')+'"'+": "+'"'+PATH_IMG+f"hulls/Tier {round(ws.cell(HULLS_OFFSET+2, i).value)}/"+name.replace("\n", "\\n").replace('"', '\\"').replace("/", "").replace('"', '').replace("*", "")+" Plain.png"+'"'
    i += 1
    
i = 2
while ws.cell(TURRETS_OFFSET+0, i).value is not None:
    if i != 2:
        turrets_string += ", \n"
    print(ws.cell(TURRETS_OFFSET+0, i).value)
    
    name = ws.cell(TURRETS_OFFSET+0, i).value\
                                        .replace("?", "")\
                                        .replace(" (!)", "")
    
    turrets_string += '"'+name.replace("\n", "\\n").replace('"', '\\"')+'"'+": "+'"'+PATH_IMG+f"turrets/Tier {round(ws.cell(TURRETS_OFFSET+2, i).value)}/"+name.replace("\n", "\\n").replace('"', '\\"').replace("/", "").replace('"', '').replace("*", "")+" Plain.png"+'"'
    i += 1
    
i = 2
while ws.cell(GUNS_OFFSET+0, i).value is not None:
    if i != 2:
        guns_string += ", \n"
    print(ws.cell(GUNS_OFFSET+0, i).value)
    
    name = ws.cell(GUNS_OFFSET+0, i).value\
                                    .replace("?", "")\
                                    .replace(" (!)", "")
    
    guns_string += '"'+name.replace("\n", "\\n").replace('"', '\\"')+'"'+": "+'"'+PATH_IMG+f"guns/Tier {round(ws.cell(GUNS_OFFSET+2, i).value)}/"+name.replace("\n", "\\n").replace('"', '\\"').replace("/", "").replace('"', '').replace("*", "")+".png"+'"'
    i += 1

output_string = templates.DATABASE.format(
    hulls=hulls_string,
    turrets=turrets_string,
    guns=guns_string,
    secondaries=""
)

with open(JSON_PATH, 'w',encoding='utf-8') as f:
    f.write(output_string)