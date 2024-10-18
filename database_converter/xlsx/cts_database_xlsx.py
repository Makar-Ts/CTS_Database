# pylint: disable=line-too-long, invalid-name, import-error, multiple-imports, unspecified-encoding, broad-exception-caught, trailing-whitespace, no-name-in-module, unused-import



# ---------------------------------------------------------------------------- #
#                               Get Database Name                              #
# ---------------------------------------------------------------------------- #


import time

DATABASE_NAME = input()
start_time = time.time()



# ---------------------------------------------------------------------------- #
#                               Libraries Import                               #
# ---------------------------------------------------------------------------- #


from openpyxl import load_workbook
import sys, os.path
import re
import codecs



parent_dir = os.path.dirname(os.path.realpath(__file__ + "/../"))
sys.path.append(parent_dir)
import modules_templates as templates



# ---------------------------------------------------------------------------- #
#                                Parcing Prepare                               #
# ---------------------------------------------------------------------------- #


DATABASE_PATH = os.path.join(sys.path[0], DATABASE_NAME)
JSON_PATH = os.path.join(sys.path[0], "database.json")

wb = load_workbook(filename=DATABASE_PATH, read_only=True, data_only=True)
ws = wb['Raw Data']


HULLS_OFFSET = 1    #Index offset
TURRETS_OFFSET = 28
GUNS_OFFSET = 50
SEC_OFFSET = 69

hulls_string = ""
turrets_string = ""
guns_string = ""
sec_string = ""

hulls_converting_time = 0
turrets_converting_time = 0
guns_converting_time = 0
sec_converting_time = 0



# ---------------------------------------------------------------------------- #
#                                   Fuctions                                   #
# ---------------------------------------------------------------------------- #


def as_string(a): return '"'+a.replace('"', '\\"')+'"' if a else '""'


def obtain_parcing(obtain):
    """
    This function takes an obtain string and parses it to extract the "Obtain" data.

    Parameters:
    - obtain (str): The string containing the obtain information.

    Returns:
    - A tuple containing the following:
        - A string with the updated obtain information after parsing.
        - A dictionary containing the resources required for the item.
        - A list containing the module type and module name of the item, if applicable.
    """
    
    add_obtain = ""
    obtain_str_prev = obtain

    arr = obtain_str_prev.split("\n")[1:]
    for j in arr:
        if j == "":
            arr.remove(j)
            
    
    # --------------------------------- Requires --------------------------------- #

    requires_module_type = "None"
    requires_module_name = ""

    if len(arr) > 0 and "Requires" in arr[0]:
        dnb = arr[0].replace("Requires ", "").split(" [")
        requires_module_type = dnb[1].replace("]", "").lower()+"s"
        requires_module_name = dnb[0]

        obtain_str_prev = obtain_str_prev.split("\nRequires")[0]
        if "Requires" in obtain_str_prev:
            obtain_str_prev = obtain_str_prev.split("Requires")[0]

        print("├ requires:", requires_module_type, requires_module_name)
        print("├ obtain_str:", obtain_str_prev)

        arr = arr[1:]

    requires = [requires_module_type, requires_module_name]
    
    
    # --------------------------- Offsale or Daily item -------------------------- #

    if "Offsale" in obtain:
        add_obtain = ", Offsale"
        obtain_str_prev = obtain_str_prev.replace("Offsale\n", "")
    if "⭐" in obtain: # In Daily Shop
        add_obtain = " | "+obtain.split("\n")[-1]
        obtain_str_prev = obtain_str_prev.replace("\n"+obtain.split("\n")[-1], "")

    
    # -------------------- Blueprints, Built-into ot Joe Shack ------------------- #
    
    if "Blueprints" in obtain_str_prev:
        obtain_str = "Blueprints"+add_obtain
        resources = '{"' + ', "'.join(map(lambda x: x.replace(":", '":'), arr)) +'}'
    elif "Built-into" in obtain_str_prev:
        obtain_str = obtain_str_prev.split("Built-into")[0]+add_obtain
        resources = "{}"
        print("├ builtinto_str:", obtain_str)
    else: # Joe Shack
        obtain_str = obtain_str_prev+add_obtain
        resources = "{}"
        

    # ---------------------------------- Return ---------------------------------- #
    
    return obtain_str, resources, requires



# ---------------------------------------------------------------------------- #
#                                 Prepare Time                                 #
# ---------------------------------------------------------------------------- #


preapre_time = time.time()-start_time
after_time = time.time()
print("!!! Prepare time: " + str(preapre_time))



# ---------------------------------------------------------------------------- #
#                                 Hulls Parcing                                #
# ---------------------------------------------------------------------------- #


i = 2
while ws.cell(HULLS_OFFSET+0, i).value is not None:
    
    
    # ---------------------------------- Prepare --------------------------------- #
    
    if i != 2:
        hulls_string += ","
    print("┌", ws.cell(HULLS_OFFSET+0, i).value)
    print("├ index:", i)
    
    obtain_str, resources, requires = obtain_parcing(ws.cell(HULLS_OFFSET+22, i).value)
    
    
    # ----------------------------- Simple Data Parce ---------------------------- #
    
    # Armor
    armor = list(map(
                        lambda x: x.split("(")[0].replace("~", ""), 
                        ws.cell(HULLS_OFFSET+5, i).value.split("/")
                    )
                )
    if len(armor) != 3:
        print("├ armor:", armor, ws.cell(HULLS_OFFSET+0, i).value.replace("\n", "\\n").replace('"', '\\"'))
        armor = [0, 0, 0, 0]
    
    
    # Speed
    speed = ws.cell(HULLS_OFFSET+6, i).value.split("/")
    
    
    # Ammo Storage
    ammo = ws.cell(HULLS_OFFSET+10, i).value.replace("(", "").replace(")", "").replace("?", "").replace(",", ".")
    
    
    # Blowout
    blowout = -1
    if "None" in ammo:
        ammo = 0
    else:
        if "Blowout" in ammo:
            if "Partial Blowout" in ammo:
                blowout = 0
            else:
                blowout = 1
        ammo = ammo.split()[0]
    
    
    # Hull Aim
    if "No" in ws.cell(HULLS_OFFSET+11, i).value: #Hull aim
        aim = 0
    elif "Yes" in ws.cell(HULLS_OFFSET+11, i).value:
        aim = 2
    elif "Suspension Only".lower() in ws.cell(HULLS_OFFSET+11, i).value.lower():
        aim = 1
    else:
        print("├ aim:", ws.cell(HULLS_OFFSET+11, i).value, ws.cell(HULLS_OFFSET+0, i).value.replace("\n", "\\n").replace('"', '\\"'))
        aim = -1
    
    
    # Active Protection System
    aps = ws.cell(HULLS_OFFSET+20, i).value.replace("N/A", "No")
    if "Yes" in aps:
        aps = "true"
    else:
        aps = "false"
    
    
    # -------------------------------- Turretless -------------------------------- #
    
    if ws.cell(HULLS_OFFSET+12, i).value == "N/A":
        have_gun = "false"
        hull_gun = "{}"
    else:
        have_gun = "true"
        
        
        # ----------------------------- Simple Data Parce ---------------------------- #
        
        # Rangefinder type
        rangefinder = ws.cell(HULLS_OFFSET+19, i).value
        if "No" in rangefinder:
            rangefinder = "-1"
        elif "Eyeball Mk.1" in rangefinder:
            rangefinder = "0"
        elif "Stereoscopic" in rangefinder:
            rangefinder = "1"
        elif "Laser" in rangefinder:
            rangefinder = "2"
        else:
            print("├ rangefinder:", rangefinder, ws.cell(HULLS_OFFSET+0, i).value.replace("\n", "\\n").replace('"', '\\"'))
            rangefinder = "0"
            
        
        # Thermal generation
        thermal = ws.cell(HULLS_OFFSET+18, i).value.replace("N/A", "No").replace("?", "No")
        if "No" in thermal:
            thermal = "0"
        else:
            thermal = thermal.replace("Gen ", "")
        
        # Zoom max-min
        zoom_str = ws.cell(HULLS_OFFSET+15, i).value.replace("N/A", "?")
        if "?" in zoom_str:
            zoom = [-1, -1]
        else:
            zoom = zoom_str.replace("x", "").split("-")
        
        # Stabilizator
        stab = ws.cell(HULLS_OFFSET+16, i).value.replace("N/A", "No")
        if "Yes" in stab:
            stab = "true"
        else:
            stab = "false"
            
        fcs = ws.cell(HULLS_OFFSET+17, i).value.replace("?", "No").replace("N/A", "No").replace("Up to ", "").replace("km", "")
        
        
        # Reload multiplier
        reload_multi = ws.cell(HULLS_OFFSET+12, i).value
        if type(reload_multi) == str:
            reload_multi = reload_multi.replace(",", ".").split()
            if len(reload_multi) != 2:
                print("├ reload:", reload_multi, ws.cell(HULLS_OFFSET+0, i).value.replace("\n", "\\n").replace('"', '\\"'))
                reload_multi = [reload_multi[0], "0"]
        else:
            reload_multi = [reload_multi, "0"]
        
        # Limits
        limits_hor = ws.cell(HULLS_OFFSET+13, i).value.split("/")
        limits_ver = ws.cell(HULLS_OFFSET+14, i).value.split("/")
        
        
        # -------------------------------- Format Data ------------------------------- #
        
        hull_gun = templates.HULL_GUN.format(
            reload_multi=reload_multi[0],
            reload_multi_caliber=reload_multi[1].replace("(", "").replace(")", "").replace("mm", ""),
            limits_up=limits_ver[1],
            limits_down=limits_ver[0].replace("-", ""),
            limits_left=limits_hor[0].replace("-", ""),
            limits_right=limits_hor[1],
            sight_rangefinder=rangefinder,
            sight_thermal=thermal,
            sight_zoom_lower=zoom[0],
            sight_zoom_upper=zoom[1],
            stabilizer=stab,
            fcs=fcs if re.match(r'^\d+(\.\d+|)$', fcs) is not None  else -1,
        )
        
    
    # -------------------------------- Fromat Data ------------------------------- #
    
    string = templates.HULL.format(
        name=ws.cell(HULLS_OFFSET+0, i).value.replace("\n", "\\n").replace('"', '\\"'),
        description=as_string(ws.cell(HULLS_OFFSET+1, i).value),
        tier=ws.cell(HULLS_OFFSET+2, i).value,
        rarity=as_string(ws.cell(HULLS_OFFSET+3, i).value),
        obtain=as_string(obtain_str.replace("\n", "")),
        resources=resources,
        weight= ws.cell(HULLS_OFFSET+8, i).value if re.match(r'^-?\d+(?:\.\d+)*$', ws.cell(HULLS_OFFSET+8, i).value) is not None else -1,
        requires_type = as_string(requires[0]),
        requires_name = as_string(requires[1]),
        armor_front=armor[0],
        armor_back=armor[2],
        armor_side=armor[1],
        speed_acceleration=ws.cell(HULLS_OFFSET+4, i).value,
        speed_forward=speed[0],
        speed_backward=speed[1],
        speed_torque=ws.cell(HULLS_OFFSET+7, i).value.replace("k", ""),
        speed_rate=ws.cell(HULLS_OFFSET+9, i).value.replace("m", ""),
        weapon_ammo_storage=ammo,
        weapon_blowout=blowout,
        weapon_hull_aim=aim,
        weapon_aps=aps,
        weapon_have_gun=have_gun,
        weapon_gun=hull_gun,
        crew='["'+ws.cell(HULLS_OFFSET+21, i).value.replace("\n", '", "')+'"]',
        based_on=as_string(" ".join(ws.cell(HULLS_OFFSET+23, i).value.split()) if ws.cell(HULLS_OFFSET+18, i).value is not None else "None"),
        paired_gun=as_string(" ".join(ws.cell(HULLS_OFFSET+25, i).value.split()[:-1]) if ws.cell(HULLS_OFFSET+20, i).value is not None else "None"),
        paired_turret=as_string(" ".join(ws.cell(HULLS_OFFSET+24, i).value.split()[:-1]) if ws.cell(HULLS_OFFSET+19, i).value is not None else "None")
    )
    
    hulls_string += string.replace("#VALUE!", "0").replace("°", "").replace("�", "").replace(" (!)", "")

    
    # ---------------------------------- Finish ---------------------------------- #
    
    i += 1
    print("└ Complete")


# -------------------------------- Statistics -------------------------------- #

hulls_converting_time = round(time.time()-start_time-preapre_time, 2)
hulls_count = i
print("!!! Hulls converting time: " + str(hulls_converting_time))    



# ---------------------------------------------------------------------------- #
#                                Turrets Parcing                               #
# ---------------------------------------------------------------------------- #


i = 2
while ws.cell(TURRETS_OFFSET+0, i).value is not None:
    
    
    # ---------------------------------- Prepare --------------------------------- #
    
    if i != 2:
        turrets_string += ","
    print("┌", ws.cell(TURRETS_OFFSET+0, i).value)
    print("├ index:", i)
    
    obtain_str, resources, requires = obtain_parcing(ws.cell(TURRETS_OFFSET+17, i).value)
    
    
    # ----------------------------- Simple Data Parce ---------------------------- #
    
    # Ammo Storage
    ammo = ws.cell(TURRETS_OFFSET+9, i).value.replace("(", "").replace(")", "").replace("?", "").replace(",", ".")
    
    
    # Blowout and Clip
    blowout = -1
    clip = 0
    
    if "None" in ammo:
        ammo = 0
    else:
        if "Blowout" in ammo:
            if "Partial Blowout" in ammo:
                blowout = 0
            else:
                blowout = 1
            
        if "Ready Rack" in ammo:
            clip = 1
        ammo = ammo.split()[0]
    
    
    # Armor
    armor = list(map(lambda x: x.split("(")[0].replace("~", ""), ws.cell(TURRETS_OFFSET+4, i).value.split("/")))
    if len(armor) != 3:
        print("├ armor:",armor, ws.cell(TURRETS_OFFSET, i).value.replace("\n", "\\n").replace('"', '\\"'))
        armor = [0, 0, 0, 0]
    
    
    # Active Protection System
    aps = ws.cell(TURRETS_OFFSET+15, i).value.replace("N/A", "No")
    if "Yes" in aps:
        aps = "true"
    else:
        aps = "false"
    
    
    # Thermal generation
    thermal = ws.cell(TURRETS_OFFSET+12, i).value.replace("N/A", "No").replace("?", "No")
    if "No" in thermal:
        thermal = "0"
    else:
        thermal = thermal.replace("Gen ", "")
    
    
    # Zoom max-min
    zoom_str = ws.cell(TURRETS_OFFSET+13, i).value.replace("N/A", "?")
    if "?" in zoom_str:
        zoom = [-1, -1]
    else:
        zoom = zoom_str.replace("x", "").split("-")
    
    
    # Stabilizator
    stab = ws.cell(TURRETS_OFFSET+5, i).value.replace("N/A", "No")
    if "Yes" in stab:
        stab = "true"
    else:
        stab = "false"
    
    
    # Rangefinder type
    rangefinder = ws.cell(TURRETS_OFFSET+14, i).value
    if "No" in rangefinder:
        rangefinder = "-1"
    elif "Eyeball Mk.1" in rangefinder:
        rangefinder = "0"
    elif "Stereoscopic" in rangefinder:
        rangefinder = "1"
    elif "Laser" in rangefinder:
        rangefinder = "2"
    else:
        print("├ rangefinder:", rangefinder, ws.cell(TURRETS_OFFSET+14, i).value.replace("\n", "\\n").replace('"', '\\"'))
        rangefinder = "0"
    
    
    # FCS distance in km
    fcs = ws.cell(TURRETS_OFFSET+11, i).value.replace("?", "No").replace("N/A", "No").replace("Up to ", "").replace("km", "")
    
    
    # Reload Multiplier
    reload_multi = ws.cell(TURRETS_OFFSET+10, i).value.replace(",", ".").split()
    if len(reload_multi) != 2:
        reload_multi = [reload_multi[0], "0"]

    
    # Speed and Limits
    limits_ver = ws.cell(TURRETS_OFFSET+7, i).value.split("/")
    speed = ws.cell(TURRETS_OFFSET+6, i).value.replace(" ", "").split("/")
    
    
    # -------------------------------- Format Data ------------------------------- #
    
    string = templates.TURRET.format(
        name=ws.cell(TURRETS_OFFSET, i).value.replace("\n", "\\n").replace('"', '\\"'),
        description=as_string(ws.cell(TURRETS_OFFSET+1, i).value),
        tier=ws.cell(TURRETS_OFFSET+2, i).value,
        rarity=as_string(ws.cell(TURRETS_OFFSET+3, i).value),
        obtain=as_string(obtain_str.replace("\n", "")),
        resources=resources,
        weight= ws.cell(TURRETS_OFFSET+8, i).value if re.match(r'^-?\d+(?:\.\d+)*$', ws.cell(TURRETS_OFFSET+8, i).value) is not None else -1,
        requires_type = as_string(requires[0]),
        requires_name = as_string(requires[1]),
        armor_front=armor[0],
        armor_back=armor[2],
        armor_side=armor[1],
        weapon_ammo_storage=ammo,
        weapon_blowout=blowout,
        weapon_clip=clip,
        weapon_aps=aps,
        weapon_fcs=fcs if re.match(r'^\d+(\.\d+|)$', fcs) is not None  else -1,
        weapon_stabilizer=stab,
        weapon_sight_thermal = thermal,
        weapon_sight_rangefinder=rangefinder,
        weapon_zoom_lower = zoom[0],
        weapon_zoom_upper = zoom[1],
        weapon_gun_reload_multi=reload_multi[0],
        weapon_gun_reload_multi_caliber=reload_multi[1].replace("(", "").replace(")", "").replace("mm", ""),
        weapon_gun_limits_up=limits_ver[1],
        weapon_gun_limits_down=limits_ver[0].replace("-", ""),
        weapon_gun_speed_vertical=speed[1][1:],
        weapon_gun_speed_horizontal=speed[0][1:],
        crew='["'+ws.cell(TURRETS_OFFSET+16, i).value.replace("\n", '", "')+'"]',
        based_on=as_string(" ".join(ws.cell(TURRETS_OFFSET+18, i).value.split()) if ws.cell(TURRETS_OFFSET+17, i).value is not None else "None"),
        paired_gun=as_string(" ".join(ws.cell(TURRETS_OFFSET+20, i).value.split()[:-1]) if ws.cell(TURRETS_OFFSET+19, i).value is not None else "None"),
        paired_hull=as_string(" ".join(ws.cell(TURRETS_OFFSET+19, i).value.split()[:-1]) if ws.cell(TURRETS_OFFSET+18, i).value is not None else "None")
    )
    
    turrets_string += string.replace("#VALUE!", "0").replace("°", "").replace(" (!)", "")
    
    
    # ---------------------------------- Finish ---------------------------------- #
    
    i += 1
    print("└ Complete")
    

# -------------------------------- Statistics -------------------------------- #

turrets_converting_time = round(time.time()-start_time-preapre_time-hulls_converting_time, 2)
turrets_count = i
print("!!! Turrets converting time: " + str(turrets_converting_time))    



# ---------------------------------------------------------------------------- #
#                                 Guns Parcing                                 #
# ---------------------------------------------------------------------------- #


i = 2
while ws.cell(GUNS_OFFSET+0, i).value is not None:
    
    
    # ---------------------------------- Prepare --------------------------------- #
    
    if i != 2:
        guns_string += ","
    print("┌", ws.cell(GUNS_OFFSET+0, i).value)
    print("├ index:", i)
    
    obtain_str, resources, requires = obtain_parcing(ws.cell(GUNS_OFFSET+14, i).value)
    
    
    # -------------------------------- Ammunition -------------------------------- #
    
    ammunition = ""
    for j in range(4):
        
        
        # ------------------------------ Get Ammo Stats ------------------------------ #
        
        splitted_ammo_data = ws.cell(GUNS_OFFSET+10+j, i).value #Ammo stats
        if splitted_ammo_data is None: break
        
        splitted_ammo_data = splitted_ammo_data.split("\n")
        
        if splitted_ammo_data[0] == "": break
        
        if j != 0: ammunition += ",\n"
        
        
        # -------------------------------- Ammo Types -------------------------------- #
        
        
        # AP, APDS, APDSFS
        if "AP" in splitted_ammo_data[0] and not "APHE" in splitted_ammo_data[0]:
            stats="""{}"""
        
        # APHE
        elif "APHE" in splitted_ammo_data[0]:
            stats=templates.AMMO_TYPES["APHE"].format(
                fuse_sensitive=splitted_ammo_data[6].split(":")[1].replace("mm", ""),
                fuse_delay=splitted_ammo_data[7].split(":")[1].replace("m", ""),
                explosive_mass=as_string(splitted_ammo_data[8].split(":")[1])
            )
        
        # HEAT and HEATFS
        elif "HEAT" in splitted_ammo_data[0]:
            stats=templates.AMMO_TYPES["HEAT"].format(
                fuse_radius=splitted_ammo_data[6].split(":")[1].replace("m", "") if len(splitted_ammo_data) > 6 else -1,
                arming_distance=splitted_ammo_data[7].split(":")[1].replace("m", "") if len(splitted_ammo_data) > 6 else -1
            )
        
        # HE and HESH
        elif "HE" in splitted_ammo_data[0] or "HESH" in splitted_ammo_data[0]:
            stats=templates.AMMO_TYPES["HE"].format(
                explosive_mass=as_string(splitted_ammo_data[6].split(":")[1]),
                fuse_radius=splitted_ammo_data[7].split(":")[1].replace("m", "") if len(splitted_ammo_data) > 7 else -1,
                arming_distance=splitted_ammo_data[8].split(":")[1].replace("m", "") if len(splitted_ammo_data) > 7 else -1
            )
        
        # ATGM
        elif "ATGM" in splitted_ammo_data[0]:
            stats=templates.AMMO_TYPES["ATGM"].format(
                range=splitted_ammo_data[6].split(":")[1].replace("km", ""),
                fuse_radius=splitted_ammo_data[7].split(":")[1].replace("m", "") if len(splitted_ammo_data) > 7 else -1,
                arming_distance=splitted_ammo_data[8].split(":")[1].replace("m", "") if len(splitted_ammo_data) > 7 else -1
            )
        else:
            stats="""{}"""
        
        
        # -------------------------------- Format Data ------------------------------- #
        
        ammunition += templates.AMMUNITION.format(
            type=as_string(splitted_ammo_data[0]),
            penetration_0deg=splitted_ammo_data[1].split(":")[1].replace("mm", ""),
            penetration_30deg=splitted_ammo_data[2].split(":")[1].replace("mm", ""),
            penetration_60deg=splitted_ammo_data[3].split(":")[1].replace("mm", ""),
            velocity=splitted_ammo_data[4].split(":")[1].replace("m/s", ""),
            ricochet_angle=splitted_ammo_data[5].split(":")[1],
            stats=stats
        )
            
    
    # -------------------------------- Format Data ------------------------------- #
    
    string = templates.GUNS.format(
        name=ws.cell(GUNS_OFFSET, i).value.replace("\n", "\\n").replace('"', '\\"'),
        description=as_string(ws.cell(GUNS_OFFSET+1, i).value),
        tier=ws.cell(GUNS_OFFSET+2, i).value,
        rarity=as_string(ws.cell(GUNS_OFFSET+3, i).value),
        obtain=as_string(obtain_str.replace("\n", "")),
        resources=resources,
        weight= ws.cell(GUNS_OFFSET+8, i).value if re.match(r'^-?\d+(?:\.\d+)*$', ws.cell(GUNS_OFFSET+8, i).value) is not None else -1,
        requires_type = as_string(requires[0]),
        requires_name = as_string(requires[1]),
        weapon_clip=ws.cell(GUNS_OFFSET+9, i).value if type(ws.cell(GUNS_OFFSET+9, i).value) == int else 0,
        weapon_reload=ws.cell(GUNS_OFFSET+7, i).value.replace("s", "") if re.match(r'^-?\d+(?:\.\d+)*$', ws.cell(GUNS_OFFSET+7, i).value.replace("s", "")) else -1,
        weapon_accuracy=ws.cell(GUNS_OFFSET+4, i).value if ws.cell(GUNS_OFFSET+4, i).value.isdigit() else -1,
        weapon_ammo_volume=ws.cell(GUNS_OFFSET+5, i).value if re.match(r'^-?\d+(?:\.\d+)*$', ws.cell(GUNS_OFFSET+5, i).value) else -1,
        weapon_caliber=ws.cell(GUNS_OFFSET+6, i).value.replace("mm", ""),
        weapon_ammunition=ammunition,
        based_on=as_string(" ".join(ws.cell(GUNS_OFFSET+15, i).value.split()) if "None" not in ws.cell(GUNS_OFFSET+15, i).value else ""),
        paired_turret=as_string(" ".join(ws.cell(GUNS_OFFSET+16, i).value.split()[:-1]) if "None" not in ws.cell(GUNS_OFFSET+16, i).value else ""),
        paired_hull=as_string(" ".join(ws.cell(GUNS_OFFSET+17, i).value.split()[:-1]) if "None" not in ws.cell(GUNS_OFFSET+17, i).value else "")
    )
    
    guns_string += string.replace("#VALUE!", "0").replace("°", "").replace(" (!)", "")
    
    
    # ---------------------------------- Finish ---------------------------------- #
    
    i += 1
    print("└ Complete")


# -------------------------------- Statistics -------------------------------- #

guns_converting_time = round(time.time()-start_time-preapre_time-hulls_converting_time-turrets_converting_time, 2)
guns_count = i
print("!!! Guns converting time: " + str(guns_converting_time))



# ---------------------------------------------------------------------------- #
#                               Secondary Parcing                              #
# ---------------------------------------------------------------------------- #


i = 2
while ws.cell(SEC_OFFSET+0, i).value is not None:
    
    
    # ---------------------------------- Prepare --------------------------------- #
    
    if i != 2:
        sec_string += ","
    print("┌", ws.cell(SEC_OFFSET+0, i).value)
    print("├ index:", i)
    
    
    # ------------------------------ Secondary Parce ----------------------------- #
    
    ammunition = ""
    for j in range(7):
        
        
        # ---------------------------------- Prepare --------------------------------- #
        
        ammo_data = ws.cell(SEC_OFFSET+3+j, i).value #Ammo stats
        if ammo_data is None: break
        
        splitted_ammo_data = ammo_data.split("\n")
        
        if splitted_ammo_data[0] == "": break
        
        if j != 0: ammunition += ",\n"
        
        is_ATGM = False
        
        
        # ----------------------------------- Type ----------------------------------- #
        
        # AP, APDS, APDSFS
        if "AP" in splitted_ammo_data[0] and not "APHE" in splitted_ammo_data[0]:
            stats="""{}"""
        
        # APHE
        elif "APHE" in splitted_ammo_data[0]:
            stats=templates.AMMO_TYPES["APHE"].format(
                fuse_sensitive=splitted_ammo_data[6].split(":")[1].replace("mm", ""),
                fuse_delay=splitted_ammo_data[7].split(":")[1].replace("m", ""),
                explosive_mass=as_string(splitted_ammo_data[8].split(":")[1])
            )
            
        # HEAT and HEATFS
        elif "HEAT" in splitted_ammo_data[0]:
            stats=templates.AMMO_TYPES["HEAT"].format(
                fuse_radius=splitted_ammo_data[6].split(":")[1].replace("m", "") if len(splitted_ammo_data) > 7 else -1,
                arming_distance=splitted_ammo_data[7].split(":")[1].replace("m", "") if len(splitted_ammo_data) > 7 else -1
            )
        
        # HE and HESH
        elif "HE" in splitted_ammo_data[0] or "HESH" in splitted_ammo_data[0]:
            stats=templates.AMMO_TYPES["HE"].format(
                explosive_mass=as_string(splitted_ammo_data[6].split(":")[1]),
                fuse_radius=splitted_ammo_data[7].split(":")[1].replace("m", "") if len(splitted_ammo_data) > 8 else -1,
                arming_distance=splitted_ammo_data[8].split(":")[1].replace("m", "") if len(splitted_ammo_data) > 8 else -1
            )
        
        # ATGM
        elif "ATGM" in splitted_ammo_data[0]:
            is_ATGM = True
            
            
            # ATGM OTA or FS
            if len(splitted_ammo_data) > 8:
                if "Fuse Radius" in splitted_ammo_data[7].split(":")[0]:
                    fuse_radius     = splitted_ammo_data[7].split(":")[1].replace("m", "")
                    arming_distance = splitted_ammo_data[8].split(":")[1].replace("m", "")
                else:
                    fuse_radius     = -1
                    arming_distance = -1
            else:
                fuse_radius     = -1
                arming_distance = -1
            
            
            reload = splitted_ammo_data[-3].split(":")[1].replace("s", "")
            
            ammo_count_splitted = splitted_ammo_data[-2].split()
            reload_count = ammo_count_splitted[-2].replace("(+", "")
            reload_count = reload_count if re.match(r'^-?\d+(?:\.\d+)*$', reload_count) is not None else 0
            
            ammo_count = ammo_count_splitted[-3]
            
            max_launch_speed = splitted_ammo_data[-1].split(":")[1]
            
            
            stats=templates.AMMO_TYPES["ATGM"].format(
                range=splitted_ammo_data[6].split(":")[1].replace("km", ""),
                fuse_radius=fuse_radius,
                arming_distance=arming_distance
            )
        else:
            stats="""{}"""
        
        
        # --------------------------------- Position --------------------------------- #
        
        types = splitted_ammo_data[0].split()
        if types[0] in ("Coax", "Roof", "Hull"):
            att_type  = types[0]
            ammo_type = types[2]
            caliber   = types[1].replace("mm", "")
        elif "ATGM" in types[0]:
            att_type  = "Coax"
            ammo_type = splitted_ammo_data[0]
            caliber   = -1
        elif re.match(r'^-?\d+(?:\.\d+)*$', types[0].replace("mm", "")) is not None:
            att_type  = "Coax"
            ammo_type = types[1]
            caliber   = types[0].replace("mm", "")
        else:
            print("├ Unknown secondary type: " + types[0])
            continue
        
        
        # ------------------------------- Reloads Count ------------------------------ #
        
        if not is_ATGM:
            ammo_count_splitted = splitted_ammo_data[-1].split()
            if "reloads" in ammo_count_splitted[-1]:
                reload_count = ammo_count_splitted[-2].replace("(+", "")
                reload_count = reload_count if re.match(r'^-?\d+(?:\.\d+)*$', reload_count) is not None else 0
                
                ammo_count = ammo_count_splitted[-3]
            else:
                reload_count = 0
                ammo_count = ammo_count_splitted[-1]
            
        if re.match(r'^-?\d+(?:\.\d+)*$', ammo_count) is None:
            print("├ Invalid ammo count: " + ammo_count)
            ammo_count = 0

        
        # -------------------------------- Format Data ------------------------------- #
        
        ammunition += templates.SECONDARIES.format(
            type=as_string(att_type),
            ammo_type=as_string(ammo_type),
            caliber=caliber,
            ammo_count=ammo_count,
            reload_count=reload_count,
            reload=reload if is_ATGM else -1,
            max_launch_speed=max_launch_speed if is_ATGM else -1,
            penetration_0deg=splitted_ammo_data[1].split(":")[1].replace("mm", ""),
            penetration_30deg=splitted_ammo_data[2].split(":")[1].replace("mm", ""),
            penetration_60deg=splitted_ammo_data[3].split(":")[1].replace("mm", ""),
            velocity=splitted_ammo_data[4].split(":")[1].replace("m/s", ""),
            ricochet_angle=splitted_ammo_data[5].split(":")[1],
            stats=stats
        )
    
    
    # -------------------------------- Format Data ------------------------------- #
        
    string = templates.SECONDARIES_ARRAY.format(
        name=ws.cell(SEC_OFFSET+0, i).value,
        module=ws.cell(SEC_OFFSET+2, i).value,
        tier=ws.cell(SEC_OFFSET+1, i).value,
        secondaries=ammunition
    )
    
    sec_string += string.replace("#VALUE!", "0").replace("°", "").replace(" (!)", "")
    
    
    # ---------------------------------- Finish ---------------------------------- #
    
    i += 1
    print("└ Complete")


# -------------------------------- Statistics -------------------------------- #

sec_converting_time = round(time.time()-start_time-preapre_time-hulls_converting_time-turrets_converting_time-guns_converting_time, 2)
sec_count = i
print("!!! Secondary converting time: " + str(sec_converting_time))



# ---------------------------------------------------------------------------- #
#                                  Write2File                                  #
# ---------------------------------------------------------------------------- #


output_string = templates.DATABASE.format(
    hulls=hulls_string,
    turrets=turrets_string,
    guns=guns_string,
    secondaries=sec_string
)
with codecs.open(JSON_PATH,mode='w',encoding='utf-8') as f:
    f.write(output_string)



# ---------------------------------------------------------------------------- #
#                                  Statistics                                  #
# ---------------------------------------------------------------------------- #


# ----------------------------------- Time ----------------------------------- #

print("\n\nConverting time: " + str(round(time.time()-start_time, 2)))
print("Hulls: " + str(hulls_converting_time))
print("Turrets: " + str(turrets_converting_time))
print("Guns: " + str(guns_converting_time))
print("Secondary: " + str(sec_converting_time))


# ---------------------------------- Amount ---------------------------------- #

print("\nModules converted: " + str(hulls_count+turrets_count+guns_count+sec_count))
print("Hulls: " + str(hulls_count) + f" ({round(hulls_converting_time/hulls_count, 3)}s per hull)")
print("Turrets: " + str(turrets_count) + f" ({round(turrets_converting_time/turrets_count, 3)}s per turret)")
print("Guns: " + str(guns_count) + f" ({round(guns_converting_time/guns_count, 3)}s per gun)")
print("Secondaries: " + str(sec_count) + f" ({round(sec_converting_time/sec_count, 3)}s per secondary)")


# --------------------------------- Database --------------------------------- #

print("Database saved: "+JSON_PATH)
