"""
Contains a SheetImageLoader class that allow you to loadimages from a sheet
"""

import io
import string
import openpyxl

from PIL import Image


class SheetImageLoader:
    """Loads all images in a sheet"""
    _images = {}

    def __init__(self, sheet):
        """Loads all sheet images"""
        sheet_images = sheet._images
        for image in sheet_images:
            row = image.anchor._from.row + 1
            col = openpyxl.utils.cell.get_column_letter(image.anchor._from.col + 1)
            # col = string.ascii_uppercase[image.anchor._from.col]
            self._images[f'{col}{row}'] = image._data

    def image_in(self, cell):
        """Checks if there's an image in specified cell"""
        return cell in self._images

    def get(self, cell):
        """Retrieves image data from a cell"""
        if cell not in self._images:
            raise ValueError("Cell {} doesn't contain an image".format(cell))
        else:
            image = io.BytesIO(self._images[cell]())
            return Image.open(image)


from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import csv, sys, os.path
import re
import codecs
import time

start_time = time.time()

parent_dir = os.path.dirname(os.path.realpath(__file__ + "/../"))
sys.path.append(parent_dir)
import modules_templates as templates

HULLS_START =  2
HULLS_OFFSET = 20
TURRETS_START =  2
TURRETS_OFFSET = 19
GUNS_START  = 2
GUNS_OFFSET = [14, 14, 15, 15, 15, 16, 15, 15, 15, 15, 16, 16]

DATABASE_NAME = input()
DATABASE_PATH = os.path.join(sys.path[0], DATABASE_NAME)
HOME_PATH = os.path.join(sys.path[0], "img")
if not os.path.exists(HOME_PATH): os.mkdir(HOME_PATH)

wb = load_workbook(filename=DATABASE_PATH)

#  HULLS PARCING
preapre_time = time.time()-start_time
after_time = time.time()
print("Prepare time: " + str(preapre_time))


if not os.path.exists(HOME_PATH): os.mkdir(HOME_PATH)

for j in range(12):
    i = 2
    
    ws = wb['HULLS']
    loader = SheetImageLoader(ws)
    Path(os.path.join(HOME_PATH, "hulls", f"Tier {j+1}")).mkdir(parents=True, exist_ok=True)
    while ws.cell(HULLS_START+HULLS_OFFSET*j-1, i).value is not None:
        if not loader.image_in(get_column_letter(i)+str(HULLS_START+HULLS_OFFSET*j)): 
            i += 2
            continue
        
        print(ws.cell(HULLS_START+HULLS_OFFSET*j-1, i).value.split("\n")[0])
        loader.get(get_column_letter(i)+str(HULLS_START+HULLS_OFFSET*j)).save(os.path.join(
                                           HOME_PATH, 
                                           "hulls", 
                                           f"Tier {j+1}", 
                                           ws.cell(HULLS_START+HULLS_OFFSET*j-1, i).value.split("\n")[0]
                                                                                    .replace('"', '')
                                                                                    .replace("/", "")
                                                                                    .replace("\\", "")
                                                                                    .replace("?", "")
                                                                                    .replace("*", "")
                                                                                    .replace(" (!)", "")+" Plain.png"))

        i += 2
        
    i = 2
    
    ws_g = wb['GUNS']
    loader_g = SheetImageLoader(ws_g)
    Path(os.path.join(HOME_PATH, "guns", f"Tier {j+1}")).mkdir(parents=True, exist_ok=True)
    oof = sum(GUNS_OFFSET[:j])
    while ws_g.cell(GUNS_START+oof-1, i).value is not None:
        if not loader_g.image_in(get_column_letter(i)+str(GUNS_START+oof)): 
            i += 2
            continue
        
        print(ws_g.cell(GUNS_START+oof-1, i).value.split("\n")[0])
        loader_g.get(get_column_letter(i)+str(GUNS_START+oof)).save(os.path.join(
                                           HOME_PATH, 
                                           "guns", 
                                           f"Tier {j+1}", 
                                           ws_g.cell(GUNS_START+oof-1, i).value.split("\n")[0]
                                                                                    .replace('"', '')
                                                                                    .replace("/", "")
                                                                                    .replace("\\", "")
                                                                                    .replace("?", "")
                                                                                    .replace("*", "")
                                                                                    .replace(" (!)", "")+".png"))
        
        i += 2
    
    i = 2
    
    ws_t = wb['TURRETS']
    loader_t = SheetImageLoader(ws_t)
    Path(os.path.join(HOME_PATH, "turrets", f"Tier {j+1}")).mkdir(parents=True, exist_ok=True)
    while ws_t.cell(TURRETS_START+TURRETS_OFFSET*j-1, i).value is not None:
        if not loader_t.image_in(get_column_letter(i)+str(TURRETS_START+TURRETS_OFFSET*j)): 
            i += 2
            continue
        
        print(ws_t.cell(TURRETS_START+TURRETS_OFFSET*j-1, i).value.split("\n")[0])
        loader_t.get(get_column_letter(i)+str(TURRETS_START+TURRETS_OFFSET*j)).save(os.path.join(
                                           HOME_PATH, 
                                           "turrets", 
                                           f"Tier {j+1}", 
                                           ws_t.cell(TURRETS_START+TURRETS_OFFSET*j-1, i).value.split("\n")[0]
                                                                                    .replace('"', '')
                                                                                    .replace("/", "")
                                                                                    .replace("\\", "")
                                                                                    .replace("?", "")
                                                                                    .replace("*", "")
                                                                                    .replace(" (!)", "")+" Plain.png"))
        
        i += 2